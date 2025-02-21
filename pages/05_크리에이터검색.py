from PIL import Image as PILImage
import streamlit as st
import pandas as pd
import traceback
import time
import json
import requests
import tempfile
import re
import io
import os
from datetime import datetime
from typing import List, Dict, Any
from io import BytesIO
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
# 스타일 관련
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


# Google API
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError



def add_image_center(ws, img_bytes, row, col):
    """
    오프셋 없이, 단순히 (row,col)에 맞춰
    셀 왼쪽 위 모서리에 이미지를 삽입.
    """
    # 1) Pillow로 이미지 열고 (필요하다면 리사이즈)
    with PILImage.open(img_bytes) as im:
        # 예: 300×300 강제 리사이즈
        im_resized = im.resize((300, 300), PILImage.LANCZOS)
        
        buffer = io.BytesIO()
        im_resized.save(buffer, format="PNG")
    buffer.seek(0)

    # 2) openpyxl Image 객체 생성
    img = ExcelImage(buffer)

    # 3) (row,col)을 셀이름(A1, B2 등)으로 환산
    col_letter = get_column_letter(col)  # ex) 1→"A", 2→"B"
    cell_name = f"{col_letter}{row}"     # ex) "A1"

    # 4) ws.add_image(img, cell_name)로 삽입
    ws.add_image(img, cell_name)

# ============================================
# 1) APIKeyManager
# ============================================
class APIKeyManager:
    def __init__(self, api_keys: List[str], daily_quota: int = 10000):
        self.api_keys = api_keys
        self.daily_quota = daily_quota
        self.current_key_index = 0
        self.current_quota_used = 0

        if not self.api_keys:
            raise ValueError("API 키 목록이 비어있습니다!")

    def get_current_key(self) -> str:
        return self.api_keys[self.current_key_index]

    def use_quota(self, cost: int):
        self.current_quota_used += cost
        if self.current_quota_used >= self.daily_quota:
            self.current_key_index += 1
            if self.current_key_index >= len(self.api_keys):
                raise RuntimeError("모든 API 키 소진!")
            self.current_quota_used = 0

# ============================================
# 2) YouTubeContentAnalyzer (키워드→쇼츠 검색)
# ============================================
class YouTubeContentAnalyzer:
    def __init__(self, api_keys: List[str], daily_quota: int = 10000):
        self.api_key_manager = APIKeyManager(api_keys, daily_quota)
        self.current_key = self.api_key_manager.get_current_key()
        self.youtube = build("youtube", "v3", developerKey=self.current_key)

        # 이미 분석한 채널 중복 방지
        self.processed_channels = set()

        # API 사용 추적
        self.api_calls = {
            "search": 0,
            "videos": 0,
            "channels": 0,
            "playlists": 0,
            "playlistItems": 0
        }
        # 대략적인 쿼터비용
        self.quota_costs = {
            "search": 100,
            "videos": 1,
            "channels": 1,
            "playlists": 1,
            "playlistItems": 1
        }

    # --- API 헬퍼 ---
    def _switch_to_next_api_key(self):
        self.api_key_manager.current_key_index += 1
        if self.api_key_manager.current_key_index >= len(self.api_key_manager.api_keys):
            st.error("[ERROR] 모든 API 키 소진. 종료.")
            raise SystemExit
        new_key = self.api_key_manager.get_current_key()
        self.youtube = build("youtube", "v3", developerKey=new_key)
        self.current_key = new_key

    def _call_api_with_retry(self, request_func, *args, **kwargs):
        max_retries = 5
        base_delay = 5
        attempt = 0
        while attempt < max_retries:
            attempt += 1
            try:
                req = request_func(*args, **kwargs)
                return req.execute()
            except HttpError as e:
                if e.resp.status == 403:
                    try:
                        err = json.loads(e.content)
                        reason = err["error"]["errors"][0]["reason"]
                    except:
                        reason = ""
                    if reason == "quotaExceeded":
                        self._switch_to_next_api_key()
                        continue
                    elif reason in ("userRateLimitExceeded","rateLimitExceeded"):
                        time.sleep(base_delay)
                        base_delay *= 2
                        continue
                    else:
                        raise
                else:
                    raise
        self._switch_to_next_api_key()
        return request_func(*args, **kwargs).execute()

    def _track_api_call(self, api_type:str):
        cost = self.quota_costs[api_type]
        self.api_key_manager.use_quota(cost)
        self.api_calls[api_type] += 1

    def get_api_usage_summary(self)->str:
        s="\n=== YouTube API 사용량 요약 ===\n"
        total=0
        for k,v in self.api_calls.items():
            cost=v*self.quota_costs[k]
            total+=cost
            s+=f"{k:15s} : {v:3d} calls (quota cost:{cost:4d})\n"
        s+=f"\n총 quota 사용량: {total}"
        return s

    # =======================
    # (A) 키워드별 쇼츠 검색
    # =======================
    def search_shorts_for_keyword(self, keyword:str, max_results:int=50)->pd.DataFrame:
        """
        1) 검색(type=video, videoDuration=short, order=viewCount), max_results=50
        2) 상위30
        3) 채널 subscriberCount >=5만 → 그 채널 쇼츠 30개 분석
        4) 결과(대표콘텐츠=검색된쇼츠) DF
        """
        # search
        self._track_api_call("search")
        search_resp = self._call_api_with_retry(
            self.youtube.search().list,
            part="snippet",
            q=keyword,
            type="video",
            videoDuration="short",
            order="viewCount",
            maxResults=max_results
        )
        items = search_resp.get("items",[])
        if not items:
            return pd.DataFrame()

        video_ids=[ it["id"]["videoId"] for it in items ]

        # videos
        self._track_api_call("videos")
        vids_resp = self._call_api_with_retry(
            self.youtube.videos().list,
            part="snippet,statistics,contentDetails",
            id=",".join(video_ids)
        )
        v_items=vids_resp.get("items",[])
        if not v_items:
            return pd.DataFrame()

        all_shorts=[]
        for v in v_items:
            sn=v["snippet"]
            stats = v.get("statistics",{})
            vc=int(stats.get("viewCount",0))
            ch_id=sn["channelId"]
            thumb=sn["thumbnails"]["high"]["url"] if "high" in sn["thumbnails"] else ""
            all_shorts.append({
                "video_id": v["id"],
                "channel_id": ch_id,
                "title": sn["title"],
                "view_count": vc,
                "published_at": sn["publishedAt"],
                "thumbnail_url": thumb,
                "keyword": keyword
            })
        # 조회수순 정렬
        all_shorts.sort(key=lambda x:x["view_count"], reverse=True)
        top30=all_shorts[:30]
        if not top30:
            return pd.DataFrame()

        # 채널 구독자수 조회
        ch_ids=list({ s["channel_id"] for s in top30 })
        ch_map=self._fetch_channel_info(ch_ids)

        results=[]
        for s in top30:
            cid=s["channel_id"]
            if cid not in ch_map:
                continue
            if ch_map[cid]["subscriber_count"]<50000:
                continue
            if cid in self.processed_channels:
                continue

            # 쇼츠 분석
            analysis=self._analyze_channel_shorts(cid, ch_map[cid]["uploads_playlist_id"])
            row={
                "keyword": s["keyword"],
                "채널id": cid,
                "채널명": ch_map[cid]["title"],
                "채널설명": ch_map[cid]["description"],
                "구독자수": ch_map[cid]["subscriber_count"],
                "채널\n전체영상수": ch_map[cid]["video_count"],
                "채널\n전체조회수": ch_map[cid]["view_count"],
                "채널 생성일": ch_map[cid]["published_at"].split("T")[0],
                "채널링크": f"https://www.youtube.com/channel/{cid}",
                "채널로고": ch_map[cid]["channel_logo"],
                "인스타그램": ch_map[cid]["instagram_url"],
                "주간 평균\n쇼츠 업로드": "",
                "30초이상\n쇼츠 비율": "",
                "쇼츠\n평균 조회수": "",
                # 대표콘텐츠=검색된 쇼츠
                "최고조회수\n쇼츠 제목": s["title"],
                "최고조회수\n쇼츠 링크": f"https://www.youtube.com/watch?v={s['video_id']}",
                "최고조회수\n쇼츠 썸네일": s["thumbnail_url"],
                "최고조회수\n쇼츠 조회수": s["view_count"],
                "분석결과": "미달",
                "미달항목": ""

            }
            if analysis:
                over_30 = analysis["over_30s_ratio"]    # 30초 이상 쇼츠 비율 (0~1)
                avg_v   = analysis["avg_views"]         # 쇼츠 평균 조회수
                viral   = analysis["viral_shorts"]      # 10만뷰 이상 쇼츠(객체 or None)
                w_uploads = analysis["avg_weekly_uploads"]  # 주간 평균 쇼츠 업로드

                row["주간 평균\n쇼츠 업로드"]=f"{analysis['avg_weekly_uploads']:.2f}"
                row["30초이상\n쇼츠 비율"]=f"{over_30*100:.1f}%"
                row["쇼츠\n평균 조회수"]=f"{avg_v:,.0f}"

                # [● 추가] 미달 조건을 기록할 리스트
                fails = []
                
                # 1) 구독자 수 < 5만
                if ch_map[cid]["subscriber_count"] < 50000:
                    fails.append("구독자 수 미달")
                
                # 2) 30초 이상 쇼츠 비율 < 0.5
                if over_30 < 0.5:
                    fails.append("30초이상 쇼츠 비율")

                # 3) 쇼츠 평균 조회수 < 20000
                if avg_v < 20000:
                    fails.append("쇼츠 평균 조회수")

                # 4) 10만뷰 이상 쇼츠 존재 X
                if viral is None:
                    fails.append("10만뷰 이상 쇼츠 미존재")

                # 5) 주간 평균 쇼츠 업로드 < 2
                if w_uploads < 2:
                    fails.append("주간 업로드 2회 미만")
                
                # 한 항목도 미달되지 않았다면 => 통과
                if not fails:
                    row["분석결과"] = "통과"
                    row["미달항목"] = "None"
                else:
                    row["분석결과"] = "미달"
                    row["미달항목"] = ",\n ".join(fails)
            
            else:
                # [● 추가] 분석 데이터가 없는 경우(쇼츠가 거의 없는 채널 등)
                row["분석결과"] = "미달"
                row["미달항목"] = "쇼츠 데이터 부족"

            self.processed_channels.add(cid)
            results.append(row)

        if not results:
            return pd.DataFrame()
        
        df=pd.DataFrame(results)
        return df

    def _fetch_channel_info(self, channel_ids:List[str]) -> Dict[str,Dict[str,Any]]:
        ch_map={}
        if not channel_ids:
            return ch_map
        self._track_api_call("channels")
        resp = self._call_api_with_retry(
            self.youtube.channels().list,
            part="snippet,statistics,brandingSettings,contentDetails",
            id=",".join(channel_ids)
        )
        items=resp.get("items",[])
        for c in items:
            cid=c["id"]
            stats=c["statistics"]
            subs=int(stats.get("subscriberCount",0))
            sn=c["snippet"]
            brand=c.get("brandingSettings",{})
            uploads=c["contentDetails"]["relatedPlaylists"]["uploads"]
            logo=sn["thumbnails"]["high"]["url"]
            tmp={
                "channel_id": cid,
                "title": sn["title"],
                "description": sn["description"],
                "subscriber_count": subs,
                "video_count": int(stats.get("videoCount",0)),
                "view_count": int(stats.get("viewCount",0)),
                "published_at": sn["publishedAt"],
                "channel_logo": logo,
                "brandingSettings": brand,
                "uploads_playlist_id": uploads,
                "instagram_url":"None"
            }
            # 인스타
            info=self._get_channel_info(tmp)
            tmp["instagram_url"]=info["instagram_url"]
            ch_map[cid]=tmp
        return ch_map

    def _analyze_channel_shorts(self, channel_id:str, playlist_id:str)->Dict[str,Any]:
        # playlistItems → videos
        videos=[]
        page_limit=3
        cur_page=0
        next_token=None
        while True:
            if len(videos)>=30:
                break
            if cur_page>=page_limit:
                break
            cur_page+=1

            self._track_api_call("playlistItems")
            pl_resp=self._call_api_with_retry(
                self.youtube.playlistItems().list,
                part="snippet",
                playlistId=playlist_id,
                maxResults=50,
                pageToken=next_token
            )
            its=pl_resp.get("items",[])
            if not its:
                break
            v_ids=[ i["snippet"]["resourceId"]["videoId"] for i in its ]
            if not v_ids:
                break

            self._track_api_call("videos")
            vresp=self._call_api_with_retry(
                self.youtube.videos().list,
                part="contentDetails,statistics,snippet",
                id=",".join(v_ids)
            )
            for v in vresp.get("items",[]):
                if self._is_shorts(v):
                    stt=v.get("statistics",{})
                    pub=v["snippet"]["publishedAt"]
                    videos.append({
                        "id":v["id"],
                        "title":v["snippet"]["title"],
                        "published_at":pub,
                        "duration":self._convert_duration_to_seconds(v["contentDetails"]["duration"]),
                        "view_count":int(stt.get("viewCount",0)),
                        "thumbnail_url":v["snippet"]["thumbnails"]["high"]["url"]
                    })
            next_token=pl_resp.get("nextPageToken")
            if not next_token:
                break

        if not videos:
            return None
        videos.sort(key=lambda x:x["published_at"], reverse=True)
        recent=videos[:30]
        if not recent:
            return None

        # 주간 평균 업로드
        recent.sort(key=lambda x:x["published_at"])
        old_dt=datetime.strptime(recent[0]["published_at"], "%Y-%m-%dT%H:%M:%SZ")
        new_dt=datetime.strptime(recent[-1]["published_at"], "%Y-%m-%dT%H:%M:%SZ")
        days_diff=(new_dt-old_dt).days
        weeks_diff=days_diff/7.0 if days_diff>0 else 1.0
        total_count=len(recent)
        avg_weekly=total_count/weeks_diff

        # 30초 이상 비율
        over_30s = sum(1 for x in recent if x["duration"]>=30)
        over_ratio = over_30s/total_count

        # 평균 조회수(상하위10% 제외)
        vs = [ x["view_count"] for x in recent ]

        avg_v = sum(vs)/len(vs) if vs else 0

        # 10만뷰↑ 중 최대
        viral=None
        mx=0
        for x in recent:
            if x["view_count"]>100000 and x["view_count"]>mx:
                viral=x
                mx=x["view_count"]

        return {
            "weekly_uploads_consistent":False,
            "avg_weekly_uploads":avg_weekly,
            "over_30s_ratio":over_ratio,
            "avg_views":avg_v,
            "viral_shorts":viral
        }

    def _is_shorts(self, video:Dict[str,Any])->bool:
        try:
            dur=video["contentDetails"]["duration"]
            sec=self._convert_duration_to_seconds(dur)
            return (sec<=60)
        except:
            return False

    def _convert_duration_to_seconds(self, duration: str)->int:
        try:
            d=duration.replace("PT","")
            m=0;s=0
            if "M" in d:
                parts=d.split("M")
                if parts[0].isdigit():
                    m=int(parts[0])
                d=parts[1]
            if "S" in d:
                s_str=d.replace("S","")
                if s_str.isdigit():
                    s=int(s_str)
            return m*60+s
        except:
            return 0

    def _get_channel_info(self, ch_data:Dict[str,Any])->Dict[str,Any]:
        """
        snippet/statistics/brandingSettings + about page크롤링
        """
        try:
            cid=ch_data["channel_id"]
            desc=ch_data["description"]
            brand=ch_data.get("brandingSettings",{})
            instagram="None"

            # about 페이지
            about_url=f"https://www.youtube.com/channel/{cid}/about"
            try:
                r=requests.get(about_url)
                if r.status_code==200:
                    txt=r.text
                    patts=[
                        r'https?://(?:www\.)?instagram\.com/[a-zA-Z0-9_\.]+/?',
                        r'(?:https?://)?instagram\.com/[a-zA-Z0-9_\.]+/?'
                    ]
                    for p in patts:
                        matches=re.finditer(p,txt)
                        for match in matches:
                            link=match.group()
                            if not link.startswith("http"):
                                link="https://"+link
                            instagram=link
                            break
                        if instagram!="None":
                            break
            except:
                pass

            # brandingSettings customLinks
            if instagram=="None" and "channel" in brand and "customLinks" in brand["channel"]:
                c_links=brand["channel"]["customLinks"]
                for cl in c_links:
                    lurl=cl.get("url","").lower()
                    ltitle=cl.get("title","").lower()
                    if "instagram" in lurl or "instagram" in ltitle or "insta" in lurl:
                        instagram=cl["url"]
                        break

            # desc
            if instagram=="None" and desc:
                pat2=[
                    r'https?://(?:www\.)?instagram\.com/[a-zA-Z0-9_\.]+/?',
                    r'instagram\.com/[a-zA-Z0-9_\.]+/?',
                    r'인스타그램\s*[:\-]?\s*@?([a-zA-Z0-9_\.]+)(?!\.[a-z]{2,4})',
                    r'인스타\s*[:\-]?\s*@?([a-zA-Z0-9_\.]+)(?!\.[a-z]{2,4})',
                    r'instagram\s*[:\-]?\s*@?([a-zA-Z0-9_\.]+)(?!\.[a-z]{2,4})',
                    r'insta\s*[:\-]?\s*@?([a-zA-Z0-9_\.]+)(?!\.[a-z]{2,4})'
                ]
                for p in pat2:
                    matches=re.finditer(p, desc, re.IGNORECASE)
                    for match in matches:
                        if len(match.groups())>0:
                            handle=match.group(1)
                            # '.'com 등 제외
                            if re.search(r'\.(com|net|org|edu|gov|mil)$', handle.lower()):
                                continue
                            instagram=f"https://instagram.com/{handle}"
                        else:
                            raw=match.group()
                            if not raw.startswith("http"):
                                raw="https://"+raw
                            instagram=raw
                        break
                    if instagram!="None":
                        break

            return {"instagram_url":instagram}
        except:
            return {"instagram_url":"None"}


# ============================================
# 3) Streamlit 메인
# ============================================
def run_app():
    st.title("유튜브 크리에이터 검색 프로그램")

    # 1) '이미 분석 완료' 상태라면, 완료 화면만 표시하고 종료
    if st.session_state.get("analysis_done"):
        time_str = datetime.now().strftime("%Y-%m-%d-%H-%M")
        st.success("크리에이터 검색이 완료되었습니다. 아래에서 결과 파일을 다운로드할 수 있습니다.")
        st.download_button(
            label="결과 다운로드(엑셀)",
            data=st.session_state["excel_data"],
            file_name=f"youtube_channel_result_{time_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        # (선택) 다시 시작하기 버튼
        if st.button("다시 시작"):
            st.session_state.clear()  # 모든 상태 초기화
            st.rerun()
        
        st.stop()  # 이미 완료 상태이므로, 아래 로직 건너뛰기

    # 2) 아직 '분석 전' 상태라면, 업로드와 분석 버튼 표시
    uploaded_file = st.file_uploader("키워드 엑셀 업로드(.xlsx, .xls)", type=["xlsx","xls"])
    if not uploaded_file:
        st.stop()

    # 여기서 엑셀 읽고, 키워드 등 사용자 입력 처리
    try:
        df_keywords = pd.read_excel(uploaded_file)
    except Exception as e:
        st.error(f"엑셀 읽기 오류: {e}")
        st.stop()

    req_cols = {"대주제","keyword","진행상태"}
    if not req_cols.issubset(df_keywords.columns):
        st.error(f"엑셀에 {req_cols} 칼럼이 필요합니다.")
        st.stop()

    st.write(f"전체 keyword 수: {len(df_keywords)}")
    df_pending = df_keywords[df_keywords["진행상태"]=="미진행"].copy()
    st.write(f"미진행 keyword: {len(df_pending)}")

    # 1) secrets에서 API 키 목록과 daily_quota 불러오기
    #    - 만약 daily_quota 없이 keys만 저장했다면, 아래처럼 삼항연산자나 get()으로 처리 가능
    secrets_api = st.secrets["API"]
    secret_keys = secrets_api["keys"]  # ["AIzaSyAAA", "AIzaSyBBB", "AIzaSyCCC"]
    daily_quota = secrets_api.get("daily_quota", 10000)

    # 2) analyzer 생성 시에 secret_keys와 daily_quota 사용
    analyzer = YouTubeContentAnalyzer(secret_keys, daily_quota=daily_quota)

    if st.button("유튜브 쇼츠 검색/분석", key="run_button"):
        try:
            # [★ 추가] 대주제, 키워드 순으로 먼저 정렬
            df_pending.sort_values(["대주제","keyword"], inplace=True)

            grouped = df_pending.groupby("대주제")

            # (1) 채널 중복 제거용 집합 선언
            all_channel_ids = set()

            # (2) 진행 단계 계산: 일단 "키워드 수 + 엑셀 작업(5단계)"로 최소치만 세팅
            total_keywords = len(df_pending)
            processed_all = total_keywords + 5  # 우선 키워드 수 + 엑셀 5단계
            processed = 0
            prog_bar = st.progress(0)
            channel_placeholder = st.empty()

            # 결과
            results_list=[]

            for main_topic, gdf in grouped:
                # 이 대주제의 키워드들
                kw_list = gdf["keyword"].tolist()
                
                for kw in kw_list:
                    # 현재 작업 업데이트
                    channel_placeholder.text(f"작업 중: '{kw}' 키워드에 대한 쇼츠 분석 중...")
                    # 실제 검색
                    df_one = analyzer.search_shorts_for_keyword(kw)
                    if not df_one.empty:
                        channel_placeholder.text(f"작업 중: '{kw}' 키워드에 대한 채널 추출 중...")
                        # '대주제' 칼럼 맨앞에 삽입
                        df_one.insert(0, "대주제", main_topic)
                        results_list.append(df_one)

                        # 진행상태=진행완료
                        mask=(df_keywords["대주제"]==main_topic)&(df_keywords["keyword"]==kw)
                        df_keywords.loc[mask,"진행상태"]="진행완료"
                    else:
                        # 검색결과 없음
                        mask=(df_keywords["대주제"]==main_topic)&(df_keywords["keyword"]==kw)
                        df_keywords.loc[mask,"진행상태"]="검색결과 없음"

                    # (1) 이 키워드에서 나온 "채널id"들을 전부 all_channel_ids에 추가
                    if not df_one.empty:
                        unique_chs = df_one["채널id"].unique()
                        all_channel_ids.update(unique_chs)

                    # (2) 키워드 1개 검색·분석 끝났으므로, 1 증가
                    processed+=1
                    prog_bar.progress(int(processed/processed_all*100))


            # (1) 모든 키워드 처리 끝
            channel_placeholder.text("모든 키워드 검색 완료.")

            # (2) "분석된 채널" 개수 반영
            channel_count = len(all_channel_ids)
            processed_all += channel_count   # 채널 수 만큼 전체 스텝을 추가
            processed += channel_count       # 이미 channel analysis는 끝났으므로, 지금 한꺼번에 처리했다고 반영
            prog_bar.progress(int(processed / processed_all * 100))

            # (3) 이후 기존 로직
            channel_placeholder.text("작업 중: 엑셀 변환 준비...")


            if results_list:
                final_df = pd.concat(results_list, ignore_index=True)
            
                # 혹시 대주제/키워드에 숨어있는 공백/줄바꿈 제거
                final_df["대주제"] = final_df["대주제"].astype(str).str.strip().str.replace("\n","").str.replace("\r","")
                final_df["keyword"] = final_df["keyword"].astype(str).str.strip().str.replace("\n","").str.replace("\r","")

                # 만약 숨은 유니코드(Zero-width space 등) 의심된다면 추가 처리
                # 예) \u200b, \u200c 등 제거
                import re
                zero_width_chars = r'[\u200B-\u200F\uFEFF]' 
                final_df["대주제"] = final_df["대주제"].apply(lambda x: re.sub(zero_width_chars, '', x))
                final_df["keyword"] = final_df["keyword"].apply(lambda x: re.sub(zero_width_chars, '', x))

                for val in final_df["대주제"].unique():
                    print(repr(val), [ord(ch) for ch in val])

                # “분석결과”도 혹시 모를 공백 제거
                final_df["분석결과"] = final_df["분석결과"].astype(str).str.strip()

                # 예) 특정 대주제='라이프스타일', 키워드='인테리어' 로 의심
                temp = final_df[
                    (final_df["대주제"] == "라이프스타일") &
                    (final_df["keyword"] == "인테리어")
                ]

                print(temp[["대주제","keyword","분석결과"]])
                print(temp.sort_values("분석결과"))  # "통과"가 정말 아래에 오는지
                print(final_df.groupby(["대주제","keyword","분석결과"]).size())


                final_df["분석결과_순위"] = final_df["분석결과"].map({"통과":0, "미달":1})
                final_df.sort_values(["대주제","keyword","분석결과_순위"], inplace=True)

                final_df.reset_index(drop=True, inplace=True)
                print(final_df[["대주제","keyword","분석결과"]].head(30))

                # 먼저 NA 여부나 기타 로직
                print(final_df["분석결과_순위"].isna().sum())

                # 그 다음에 컬럼 제거
                final_df.drop(columns=["분석결과_순위"], inplace=True)
            else:
                final_df = pd.DataFrame()

            # 원하는 순서를 리스트로 명시 (예시)
            desired_order = [
                "대주제",
                "keyword",
                "채널명",
                "구독자수",
                "분석결과",
                "미달항목",
                "쇼츠\n평균 조회수",
                "주간 평균\n쇼츠 업로드",
                "30초이상\n쇼츠 비율",
                "최고조회수\n쇼츠 제목",
                "최고조회수\n쇼츠 조회수",
                "최고조회수\n쇼츠 썸네일",
                "최고조회수\n쇼츠 링크",
                "채널설명",
                "채널로고",
                "채널링크",
                "인스타그램",
                "채널\n전체영상수",
                "채널\n전체조회수",
                "채널 생성일",
                "채널id"
            ]


            # 실제로 열 순서를 재배치
            final_df = final_df[desired_order]

            # 엑셀에 이미지 삽입
            if not final_df.empty:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title="채널분석결과"
                
                headers=list(final_df.columns)

                # 1) 스타일 객체 정의
                header_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # 연두색 배경
                header_font = Font(bold=True)  # 볼드체
                header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

                data_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
                data_align_description = Alignment(horizontal="center", vertical="center", wrap_text=True)

                light_green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                light_red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                bold_font       = Font(bold=True)
                bold_big_font   = Font(bold=True, size=15)
                bold_mid_font = Font(bold=True, size=13)
                default_font    = Font(bold=False)

                thin_border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )

                thin_border_gray = Border(
                    left=Side(style='thin', color='E2E2E2'),
                    right=Side(style='thin', color='E2E2E2'),
                    top=Side(style='thin', color='E2E2E2'),
                    bottom=Side(style='thin', color='E2E2E2')
                )

                processed+=1
                prog_bar.progress(int(processed/processed_all*100))

                # 2) 헤더 행 쓰기
                for col_idx, col_name in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_align
                    cell.border = thin_border

                # 3) 데이터 행 쓰기
                #    - "채널로고", "최고조회수 쇼츠 썸네일" → cell.value = ""
                #    - 링크 칼럼(채널링크, 최고조회수 쇼츠 링크, 인스타그램 등)은 하이퍼링크 적용
                link_columns = {"채널링크", "최고조회수\n쇼츠 링크", "인스타그램"}  # 필요시 조정
                logo_columns = {"채널로고", "최고조회수\n쇼츠 썸네일"}
                description_columns = {"채널설명", "최고조회수\n쇼츠 제목"}

                for row_i, row_data in final_df.iterrows():
                    excel_row = row_i + 2
                    for col_idx, col_name in enumerate(headers, start=1):
                        raw_value = row_data[col_name]
                        cell = ws.cell(row=excel_row, column=col_idx)
                        
                        # 기본: 텍스트로 저장 + 가운데 정렬
                        cell.value = raw_value
                        cell.alignment = data_align
                        
                        # ---- [A] 칼럼별 "볼드 처리" ----
                        if col_name in [
                            "채널명",        # 글씨크기 13 + 볼드
                            "대주제", 
                            "keyword", 
                            "분석결과", 
                            "미달항목", 
                            "구독자수", 
                            "채널 생성일", 
                            "주간 평균\n쇼츠 업로드", 
                            "30초이상\n쇼츠 비율", 
                            "쇼츠\n평균 조회수", 
                            "최고조회수\n쇼츠 조회수",
                            "채널전체영상수" 
                        ]:
                            # 특별히 "채널명"만 사이즈 13으로
                            if col_name == "채널명":
                                cell.font = bold_big_font
                            else:
                                cell.font = bold_font


                        # ---- [B] '분석결과' (통과=연초록, 미달=연빨강) ----
                        if col_name == "분석결과":
                            if raw_value == "통과":
                                cell.fill = light_green_fill
                            elif raw_value == "미달":
                                cell.fill = light_red_fill


                        # --- [C] '미달항목' (공란 -> 연초록, 값 있으면 연빨강) ---
                        if col_name == "미달항목":
                            # "None" → 초록색 글씨
                            if raw_value == "None":
                                cell.font = Font(color="0c6e0c", bold=True)  
                            else:
                                # 그 외(미달 사유 있음) → 빨강
                                cell.font = Font(color="FF0000", bold=True)


                        # ---- [D] '인스타그램' (값이 'None'이면 글씨 빨강) ----
                        if col_name == "인스타그램":
                            if raw_value == "None":
                                # 빨간색 글씨
                                cell.font = Font(color="FF0000", bold=True)  # 굵게/빨강
                            # else: 그대로 (없으면 아무 처리 안 함)


                        # ---- [E] '주간 평균 쇼츠 업로드' (값 >= 2 => 연초록, <2 => 연빨강) ----
                        if col_name == "주간 평균\n쇼츠 업로드":
                            try:
                                # 숫자로 변환
                                v = float(raw_value)
                                cell.value = v               # 숫자로 저장
                                cell.number_format = '0.00'  # 소수점 2자리 예시
                                if v >= 2.0:
                                    cell.font = Font(color="0c6e0c", bold=True, size=13)  # 굵게/초록
                                else:
                                    cell.font = Font(color="FF0000", bold=True, size=13)  # 굵게/빨강
                            except:
                                pass  # 변환 안 되면 그대로 둠


                        # ---- [F] '30초이상 쇼츠 비율' (값 >= 50% => 연초록, <50% => 연빨강) ----
                        if col_name == "30초이상\n쇼츠 비율":
                            # 예: "65.3%" 형태라고 가정
                            val_str = str(raw_value).replace("%","").strip()
                            try:
                                ratio = float(val_str)
                                # 엑셀에 실제 퍼센트로 저장하려면 (0.653)으로 저장하고 number_format = '0.0%'
                                # 혹은 그냥 65.3 (정수/실수)로 저장하고 수식 없이도 가능
                                cell.value = ratio / 100
                                cell.number_format = '0.0%'  # x.x%
                                
                                if ratio >= 50.0:
                                    cell.font = Font(color="0c6e0c", bold=True, size=13)  # 굵게/초록
                                else:
                                    cell.font = Font(color="FF0000", bold=True, size=13)  # 굵게/빨강
                            except:
                                pass


                        # ---- [G] '쇼츠 평균 조회수' (값 >=2만 => 연초록, <2만 => 연빨강) ----
                        if col_name == "쇼츠\n평균 조회수":
                            try:
                                v = float(str(raw_value).replace(",",""))  # 혹시 콤마가 있을 경우 제거
                                cell.value = v
                                cell.number_format = '#,##0'  # 천단위 콤마
                                
                                if v >= 20000:
                                    cell.font = Font(color="0c6e0c", bold=True, size=13)  # 굵게/초록
                                else:
                                    cell.font = Font(color="FF0000", bold=True, size=13)  # 굵게/빨강
                            except:
                                pass


                        # ---- [H] '구독자수', '최고조회수 쇼츠 조회수', '채널전체영상수' 등 콤마 처리 ----
                        #     (이미 "쇼츠 평균 조회수"는 위에서 처리했으므로 생략)
                        if col_name in ["구독자수", "최고조회수\n쇼츠 조회수", "채널전체영상수"]:
                            try:
                                v = float(str(raw_value).replace(",",""))
                                cell.value = v
                                cell.number_format = '#,##0'  # 천단위 콤마
                            except:
                                pass


                        # ---- [I] 채널로고/썸네일 칼럼은 텍스트 제거 ("") ----
                        if col_name in logo_columns:
                            cell.value = ""
                        

                        # ---- [J] 링크 칼럼이면 하이퍼링크 걸기 ----
                        elif col_name in link_columns:
                            # 만약 raw_value가 유효한 http URL이면
                            if isinstance(raw_value, str) and raw_value.startswith("http"):
                                cell.value = raw_value  # or "바로가기"
                                cell.hyperlink = raw_value
                                cell.style = "Hyperlink"
                                cell.alignment = data_align
                            else:
                                cell.value = str(raw_value)
                                cell.alignment = data_align


                        # ---- [K] 채널설명/최고조회수 쇼츠 제목 칼럼은 위쪽 정렬 적용 ----
                        elif col_name in description_columns:
                            cell.value = str(raw_value)
                            cell.alignment = data_align_description  # 위쪽 정렬
                        else:
                            # 일반 텍스트 칼럼
                            cell.value = str(raw_value)
                            cell.alignment = data_align

                        cell.border = thin_border_gray

                processed+=1
                prog_bar.progress(int(processed/processed_all*100))

                # 이미지 삽입
                col_logo = headers.index("채널로고") + 1
                col_thumb = headers.index("최고조회수\n쇼츠 썸네일") + 1

                for row_i, row_data in final_df.iterrows():
                    excel_row = row_i + 2

                    # 채널로고
                    logo_url = row_data["채널로고"]
                    if isinstance(logo_url, str) and logo_url.startswith("http"):
                        try:
                            r = requests.get(logo_url, timeout=5)
                            if r.status_code == 200:
                                with io.BytesIO(r.content) as buf:
                                    # 중앙 정렬 없이 바로 셀 (excel_row, col_logo)에 삽입
                                    add_image_center(ws, buf, excel_row, col_logo)
                        except Exception as e:
                            st.error(f"[ERROR] 로고 이미지 삽입 실패: {e}")

                    # 썸네일
                    thumb_url = row_data["최고조회수\n쇼츠 썸네일"]
                    if isinstance(thumb_url, str) and thumb_url.startswith("http"):
                        try:
                            r2 = requests.get(thumb_url, timeout=5)
                            if r2.status_code == 200:
                                with io.BytesIO(r2.content) as buf:
                                    # 중앙 정렬 없이 바로 셀 (excel_row, col_thumb)에 삽입
                                    add_image_center(ws, buf, excel_row, col_thumb)
                        except Exception as e:
                            st.error(f"[ERROR] 썸네일 이미지 삽입 실패: {e}")


                processed+=1
                prog_bar.progress(int(processed/processed_all*100))

                # 열너비/행높이 설정(옵션)
                ws.column_dimensions[get_column_letter(1)].width=15 # 대주제
                ws.column_dimensions[get_column_letter(2)].width=10 # keyword
                ws.column_dimensions[get_column_letter(3)].width=20 # 채널명	
                ws.column_dimensions[get_column_letter(4)].width=10 # 구독자수	
                ws.column_dimensions[get_column_letter(5)].width=10 # 분석결과
                ws.column_dimensions[get_column_letter(6)].width=20 # 미달항목 
                ws.column_dimensions[get_column_letter(7)].width=12 # 쇼츠 평균 조회수
                ws.column_dimensions[get_column_letter(8)].width=13 # 주간 평균 쇼츠 업로드
                ws.column_dimensions[get_column_letter(9)].width=10 # 30초이상 쇼츠 비율
                ws.column_dimensions[get_column_letter(10)].width=15 # 최고조회수 쇼츠 제목
                ws.column_dimensions[get_column_letter(11)].width=13 # 최고조회수 쇼츠 조회수
                ws.column_dimensions[get_column_letter(12)].width=38 # 최고조회수 쇼츠 썸네일
                ws.column_dimensions[get_column_letter(13)].width=11 # 최고조회수 쇼츠 링크
                ws.column_dimensions[get_column_letter(14)].width=25 # 채널설명
                ws.column_dimensions[get_column_letter(15)].width=38 # 채널로고	
                ws.column_dimensions[get_column_letter(16)].width=10 # 채널링크
                ws.column_dimensions[get_column_letter(17)].width=10 # 인스타그램
                ws.column_dimensions[get_column_letter(18)].width=11 # 채널 전체영상수
                ws.column_dimensions[get_column_letter(19)].width=11 # 채널 전체조회수
                ws.column_dimensions[get_column_letter(20)].width=12 # 채널 생성일	
                ws.column_dimensions[get_column_letter(21)].width=10 # 채널id

                ws.row_dimensions[1].height=40
                for r_i in range(2, ws.max_row+1):
                    ws.row_dimensions[r_i].height=230


                processed+=1
                prog_bar.progress(int(processed/processed_all*100))

                # ... (채널 분석 코드) ...
                channel_placeholder.text("모든 작업이 완료되었습니다.")  # 모든 작업 완료 메시지

                output_buffer=BytesIO()
                wb.save(output_buffer)
                output_buffer.seek(0)

                # 현재 시간 가져오기
                now = datetime.now()

                # 원하는 형식으로 문자열 변환 (YYYY-MM-DD-HH-mm)
                time_str = now.strftime("%Y-%m-%d-%H-%M")


                # 세션 상태에 "완료" + "엑셀 파일" 저장
                st.session_state["analysis_done"] = True
                st.session_state["excel_data"] = output_buffer.getvalue()
            
            else:
                st.warning("검색된 결과가 없습니다.")

        except Exception as e:
            st.error(f"오류 발생: {e}")
            traceback.print_exc()


    if "analysis_done" in st.session_state:
        time_str = datetime.now().strftime("%Y-%m-%d-%H-%M")
        
        st.success("크리에이터 검색이 완료되었습니다. 아래에서 결과 파일을 다운로드할 수 있습니다.")
        st.download_button(
            label="결과 다운로드(엑셀)",
            data=st.session_state["excel_data"],
            file_name=f"youtube_channel_result_{time_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

if __name__=="__main__":
    run_app()
